import json
import os
import re
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[1]
PACK_PATH = ROOT_DIR / "data" / "football-data-pack.js"
WORLD_CUP_URL = "https://www.football-data.co.uk/WorldCup2026.xlsx"

STANDARD_HEADER = "Div,Date,HomeTeam,AwayTeam,FTHG,FTAG,FTR,B365H,B365D,B365A"
WORLD_CUP_SHEETS = {
    "WorldCup2014": ("WORLDCUP", "2014"),
    "WorldCup2018": ("WORLDCUP", "2018"),
    "WorldCup2022": ("WORLDCUP", "2022"),
    "WorldCup2026": ("WORLDCUP", "2026"),
    "WorldCup2026Qualifiers": ("WCQ", "2026"),
}


def load_pack():
    if not PACK_PATH.exists():
        return {}
    text = PACK_PATH.read_text(encoding="utf-8")
    match = re.search(r"window\.FOOTBALL_DATA_PACK\s*=\s*(\{.*\});?\s*$", text, re.S)
    if not match:
        raise RuntimeError("football-data-pack.js 형식을 읽을 수 없습니다.")
    return json.loads(match.group(1))


def write_pack(pack):
    PACK_PATH.write_text(
        "window.FOOTBALL_DATA_PACK = "
        + json.dumps(pack, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )


def get_workbook_path():
    cache_path = Path(os.environ.get("TEMP", str(ROOT_DIR))) / "WorldCup2026.xlsx"
    if not cache_path.exists() or cache_path.stat().st_size < 1000:
        urllib.request.urlretrieve(WORLD_CUP_URL, cache_path)
    return cache_path


def normalize_date(value):
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value or "").strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", text):
        return datetime.fromisoformat(text[:10]).strftime("%d/%m/%Y")
    return text


def csv_value(value):
    text = str(value if value is not None else "").strip()
    if any(char in text for char in [",", '"', "\n", "\r"]):
        return '"' + text.replace('"', '""') + '"'
    return text


def get(row, header, names):
    for name in names:
        index = header.get(name.lower())
        if index is not None and row[index] is not None and str(row[index]).strip() != "":
            return row[index]
    return ""


def result_from_goals(home_goals, away_goals):
    try:
        home = int(home_goals)
        away = int(away_goals)
    except (TypeError, ValueError):
        return "UNKNOWN"
    if home > away:
        return "H"
    if home < away:
        return "A"
    return "D"


def convert_sheet(ws, league_key):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""
    header = {str(value or "").strip().lower(): index for index, value in enumerate(rows[0])}
    output = [STANDARD_HEADER]

    for row in rows[1:]:
        home = get(row, header, ["Home", "HomeTeam"])
        away = get(row, header, ["Away", "AwayTeam"])
        date = normalize_date(get(row, header, ["Date"]))
        home_goals = get(row, header, ["HGFT", "HG", "FTHG"])
        away_goals = get(row, header, ["AGFT", "AG", "FTAG"])
        home_odds = get(row, header, ["bet365-H", "Pinny-H", "H-Avg", "H_Avg", "H-Max", "H_Max"])
        draw_odds = get(row, header, ["bet365-D", "Pinny-D", "D-Avg", "D_Avg", "D-Max", "D_Max"])
        away_odds = get(row, header, ["bet365-A", "Pinny-A", "A-Avg", "A_Avg", "A-Max", "A_Max"])

        if not all([home, away, date, home_odds, draw_odds, away_odds]):
            continue

        result = result_from_goals(home_goals, away_goals)
        output.append(",".join(csv_value(value) for value in [
            league_key,
            date,
            home,
            away,
            home_goals,
            away_goals,
            result,
            home_odds,
            draw_odds,
            away_odds,
        ]))

    return "\n".join(output)


def main():
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else get_workbook_path()
    pack = load_pack()
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    changed = []

    for sheet_name, (league_key, season) in WORLD_CUP_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        csv_text = convert_sheet(wb[sheet_name], league_key)
        match_count = max(len(csv_text.splitlines()) - 1, 0)
        if match_count <= 0:
            continue
        pack.setdefault(league_key, {})[season] = csv_text
        changed.append((league_key, season, match_count))

    write_pack(pack)
    for league_key, season, match_count in changed:
        print(f"{league_key} {season}: {match_count}경기 추가/갱신")
    print(f"완료: {len(changed)}개 시즌 갱신")


if __name__ == "__main__":
    main()
